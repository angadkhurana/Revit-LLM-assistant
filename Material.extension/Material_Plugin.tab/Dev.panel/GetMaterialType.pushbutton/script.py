#! python3
__title__ = 'wassup'
__author__  = 'angad'
__doc__ = """random"""


from Autodesk.Revit.DB import *

import subprocess
from pyrevit import  DB
import openpyxl
import xlwings as xw
from pyxlsb import open_workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

uidoc = __revit__.ActiveUIDocument
doc = __revit__.ActiveUIDocument.Document
from snippets.selection import *


# from bs4 import BeautifulSoup 
import subprocess

if __name__ == '__main__':

    # target_environment_python = 'C:/Users/angad/anaconda3/envs/revit_llm_env/python.exe'
    # target_script = 'C:/Users/angad/OneDrive/Desktop/LangChain/Revit-LLM-assistant/qwerty.py'
    # print(call_other_script(target_environment_python, target_script, 'London'))

    material_list = ["WAL_Concrete_RC"]

    volume_of_materials_in_beams = {material:get_volume_of_material_in_component(doc, material, "beams") for material in material_list}
    volume_of_materials_in_columns = {material:get_volume_of_material_in_component(doc, material, "columns") for material in material_list}
    volume_of_materials_in_floors = {material:get_volume_of_material_in_component(doc, material, "floors") for material in material_list}
    volume_of_materials_in_foundations = {material:get_volume_of_material_in_component(doc, material, "foundations") for material in material_list}
    volume_of_materials_in_walls = {material:get_volume_of_material_in_component(doc, material, "walls") for material in material_list}



    print(volume_of_materials_in_beams)
    print(volume_of_materials_in_columns)
    print(volume_of_materials_in_floors)
    print(volume_of_materials_in_foundations)
    print(volume_of_materials_in_walls)

    # Load the original Excel file
    original_file_path = r"C:\Users\angad\OneDrive\Desktop\Walsh\original.xlsm"
    modified_file_path = r"C:\Users\angad\OneDrive\Desktop\Walsh\original3.xlsm"

    original_workbook = load_workbook(original_file_path, keep_vba=True, data_only = 'True')

    # Modify specific cells
    sheet = original_workbook['Walsh ECA Input (Gate)']  # Replace 'Sheet1' with your sheet name
    sheet['A1'] = 'Hello'
    sheet['B1'] = 'World'
    sheet['A2'] = 123
    sheet['B2'] = 456

    # Copy pictures from the original workbook to the modified workbook
    for sheet_name in original_workbook.sheetnames:
        original_sheet = original_workbook[sheet_name]
        modified_sheet = original_workbook[sheet_name]

        for img in original_sheet._images:
            modified_sheet.add_image(img)

    # Save the modified workbook
    original_workbook.save(modified_file_path)
    



    



    
   